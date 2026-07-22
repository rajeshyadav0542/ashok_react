from __future__ import annotations
import json
from pathlib import Path
from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Any, Dict, List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from user_input_parameter_api import router
from user_input_parameter_api import refresh_meta_parameters
from STRATIFICATION_V2_FAST_API import run_stratification as execute_stratification
class CampaignControlFormData(BaseModel):
    approach: str
    dataPrepFileName: str
    metaFileName: str
    mmxFileName: str
    campaign_start: str
    campaign_end: str
    pre_start: str
    pre_end: str
    balancingVariables: List[str]
    salesMetrics: List[str]
    activityTolerances: Dict[str, str]
    outlierMethod: str
    ctrl_ratio: float = 0.2
    sales_metric: str = "SALES_1"
    SEGMENT_CATEGORICAL_1: str = "null"
    SEGMENT_CATEGORICAL_2: str = "null"
    SEGMENT_CATEGORICAL_3: str = "null"
    SEGMENT_CATEGORICAL_4: str = "null"
    SEGMENT_NUMERICAL_1: str = "null"
    SEGMENT_NUMERICAL_2: str = "null"
    SEGMENT_NUMERICAL_3: str = "null"

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
@app.post("/run-stratification")
def run_stratification():
    result = execute_stratification()
    return result

from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent

PUBLIC_DIR = PROJECT_ROOT / "public"

EXCEL_PATH = PUBLIC_DIR / "Final_Results.xlsx"
UPLOAD_DIR = PUBLIC_DIR / "metadata"

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

print("PROJECT_ROOT :", PROJECT_ROOT)
print("PUBLIC_DIR   :", PUBLIC_DIR)
print("UPLOAD_DIR   :", UPLOAD_DIR)
SHEET_NAME = "CampaignControlForm"

HEADERS = [
    "submitted_at",
    "approach",
    "dataPrepFileName",
    "metaFileName",
    "mmxFileName",
    "campaign_start",
    "campaign_end",
    "pre_start",
    "pre_end",
    "balancingVariables",
    "salesMetrics",
    "activityTolerances",
    "outlierMethod",
    "ctrl_ratio",
    "sales_metric",
    "SEGMENT_CATEGORICAL_1",
    "SEGMENT_CATEGORICAL_2",
    "SEGMENT_CATEGORICAL_3",
    "SEGMENT_CATEGORICAL_4",
    "SEGMENT_NUMERICAL_1",
    "SEGMENT_NUMERICAL_2",
    "SEGMENT_NUMERICAL_3",
]
print("Upload Folder:", UPLOAD_DIR.resolve())

def get_workbook() -> Workbook:
    if EXCEL_PATH.exists():
        try:
            return load_workbook(EXCEL_PATH)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Unable to read Excel file: {exc}")

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.save(EXCEL_PATH)
    return workbook


def ensure_sheet(workbook: Workbook):
    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
    else:
        sheet = workbook.create_sheet(SHEET_NAME)
        sheet.append(HEADERS)
    if sheet.max_row == 1 and [cell.value for cell in sheet[1]] != HEADERS:
        sheet.delete_rows(1)
        sheet.append(HEADERS)
    return sheet


def safe_save_workbook(workbook: Workbook, path: Path):
    temp_path = path.with_name(f"{path.stem}.tmp{path.suffix}")
    try:
        workbook.save(temp_path)
        temp_path.replace(path)
    except PermissionError as exc:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise HTTPException(
            status_code=423,
            detail=(
                "Unable to save Excel file: file is locked or open in another application. "
                "Close the workbook in Excel or stop any process holding the file, then retry."
            ),
        ) from exc
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Unable to save Excel file: {exc}") from exc


def normalize_form_data(payload: Dict[str, Any]) -> CampaignControlFormData:
    normalized_payload: Dict[str, Any] = dict(payload)

    if "campaign_end" not in normalized_payload and "campaignEnd" in normalized_payload:
        normalized_payload["campaign_end"] = normalized_payload["campaignEnd"]
    if "pre_start" not in normalized_payload and "precampaign_start" in normalized_payload:
        normalized_payload["pre_start"] = normalized_payload["precampaign_start"]
    if "pre_end" not in normalized_payload and "preCampaignEnd" in normalized_payload:
        normalized_payload["pre_end"] = normalized_payload["preCampaignEnd"]
    if "campaign_start" not in normalized_payload and "campaignStart" in normalized_payload:
        normalized_payload["campaign_start"] = normalized_payload["campaignStart"]

    normalized_payload.setdefault("balancingVariables", [])
    normalized_payload.setdefault("salesMetrics", [])
    normalized_payload.setdefault("activityTolerances", {})
    normalized_payload.setdefault("outlierMethod", "2*SD")
    normalized_payload.setdefault("ctrl_ratio", 0.2)
    normalized_payload.setdefault("sales_metric", "SALES_1")
    normalized_payload.setdefault("SEGMENT_CATEGORICAL_1", "null")
    normalized_payload.setdefault("SEGMENT_CATEGORICAL_2", "null")
    normalized_payload.setdefault("SEGMENT_CATEGORICAL_3", "null")
    normalized_payload.setdefault("SEGMENT_CATEGORICAL_4", "null")
    normalized_payload.setdefault("SEGMENT_NUMERICAL_1", "null")
    normalized_payload.setdefault("SEGMENT_NUMERICAL_2", "null")
    normalized_payload.setdefault("SEGMENT_NUMERICAL_3", "null")

    return CampaignControlFormData(**normalized_payload)


def build_upload_path(filename: str) -> Path:
    original_name = Path(filename or "upload").name
    if not original_name:
        original_name = "upload"

    stem = Path(original_name).stem or "upload"
    suffix = Path(original_name).suffix
    candidate = UPLOAD_DIR / f"{stem}{suffix}"
    counter = 1

    while candidate.exists():
        candidate = UPLOAD_DIR / f"{stem}_{counter}{suffix}"
        counter += 1

    return candidate


async def save_uploaded_file(upload: Optional[UploadFile]) -> Optional[str]:
    if upload is None or upload.filename is None:
        return None

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    destination = build_upload_path(upload.filename)

    print(f"Saving file: {upload.filename}")
    print(f"Destination: {destination.resolve()}")

    try:
        with destination.open("wb") as file_handle:
            while chunk := await upload.read(1024 * 1024):
                file_handle.write(chunk)

        print("Saved successfully.")

    except Exception as e:
        print("Upload Error:", e)
        raise HTTPException(status_code=500, detail=str(e))

    return f"/metadata/{destination.name}"


def append_submission(sheet, submission: CampaignControlFormData):
    values = [
        datetime.utcnow().isoformat(),
        submission.approach,
        submission.dataPrepFileName,
        submission.metaFileName,
        submission.mmxFileName,
        submission.campaign_start,
        submission.campaign_end,
        submission.pre_start,
        submission.pre_end,
        json.dumps(submission.balancingVariables, ensure_ascii=False),
        json.dumps(submission.salesMetrics, ensure_ascii=False),
        json.dumps(submission.activityTolerances, ensure_ascii=False),
        submission.outlierMethod,
        submission.ctrl_ratio,
        submission.sales_metric,
        submission.SEGMENT_CATEGORICAL_1,
        submission.SEGMENT_CATEGORICAL_2,
        submission.SEGMENT_CATEGORICAL_3,
        submission.SEGMENT_CATEGORICAL_4,
        submission.SEGMENT_NUMERICAL_1,
        submission.SEGMENT_NUMERICAL_2,
        submission.SEGMENT_NUMERICAL_3,
    ]
    sheet.append(values)
    for idx, _ in enumerate(values, start=1):
        sheet.column_dimensions[get_column_letter(idx)].auto_size = True


def load_submissions(sheet):
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    submissions = []
    for row in rows:
        if not row or all(value is None for value in row):
            continue

        submission = {
            "submitted_at": row[0] or "",
            "approach": row[1] or "",
            "dataPrepFileName": row[2] or "",
            "metaFileName": row[3] or "",
            "mmxFileName": row[4] or "",
            "campaign_start": row[5] or "",
            "campaign_end": row[6] or "",
            "pre_start": row[7] or "",
            "pre_end": row[8] or "",
            "balancingVariables": json.loads(row[9]) if row[9] else [],
            "salesMetrics": json.loads(row[10]) if row[10] else [],
            "activityTolerances": json.loads(row[11]) if row[11] else {},
            "outlierMethod": row[12] or "",
            "ctrl_ratio": row[13] if len(row) > 13 else 0.2,
            "sales_metric": row[14] if len(row) > 14 else "SALES_1",
            "SEGMENT_CATEGORICAL_1": row[15] if len(row) > 15 else "null",
            "SEGMENT_CATEGORICAL_2": row[16] if len(row) > 16 else "null",
            "SEGMENT_CATEGORICAL_3": row[17] if len(row) > 17 else "null",
            "SEGMENT_CATEGORICAL_4": row[18] if len(row) > 18 else "null",
            "SEGMENT_NUMERICAL_1": row[19] if len(row) > 19 else "null",
            "SEGMENT_NUMERICAL_2": row[20] if len(row) > 20 else "null",
            "SEGMENT_NUMERICAL_3": row[21] if len(row) > 21 else "null",
        }
        submissions.append(submission)
    return submissions


@app.post("/api/campaign-control")
async def save_campaign_control(request: Request):

    print("===== API HIT =====")

    form_data = await request.form()

    print("Form Keys:", list(form_data.keys()))

    payload_value = form_data["payload"]
    form_payload = json.loads(payload_value)

    normalized_submission = normalize_form_data(form_payload)
    workbook = get_workbook()
    sheet = ensure_sheet(workbook)
    append_submission(sheet, normalized_submission)
    safe_save_workbook(workbook, EXCEL_PATH)
    return {
        "success": True,
        "message": "Saved"
    }

@app.get("/api/campaign-control")
async def get_campaign_control():
    if not EXCEL_PATH.exists():
        return {"success": True, "submissions": []}

    workbook = get_workbook()
    if SHEET_NAME not in workbook.sheetnames:
        return {"success": True, "submissions": []}

    sheet = workbook[SHEET_NAME]
    submissions = load_submissions(sheet)
    return {"success": True, "submissions": submissions}

@app.post("/api/upload-file")
async def upload_file(
    dataPrepFile: UploadFile = File(None),
    metaFile: UploadFile = File(None),
    mmxFile: UploadFile = File(None),
):

    response = {}

    if dataPrepFile:
        file_path = await save_uploaded_file(dataPrepFile)
        response["dataPrepFile"] = file_path

    if metaFile:
        file_path = await save_uploaded_file(metaFile)
        refresh_meta_parameters()
        response["metaFile"] = file_path
        response["metaRefreshed"] = True

    if mmxFile:
        file_path = await save_uploaded_file(mmxFile)
        response["mmxFile"] = file_path

    if not response:
        raise HTTPException(
            status_code=400,
            detail="No file selected"
        )

    return {
        "success": True,
        **response
    }


app.include_router(router)
